from io import BytesIO
from pathlib import Path
from types import MethodType

from django.contrib import admin
from django.contrib import messages
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.contrib.auth.models import Group
from django.conf import settings
from django.http import HttpRequest
from django.http import FileResponse
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.urls import path, reverse
from django.utils.html import format_html
from openpyxl import load_workbook

from .models import ClassCatalog, MajorCatalog, MajorClassManagement, User

admin.site.site_header = "教师账号审核后台"
admin.site.site_title = "教师账号审核后台"
admin.site.index_title = "教师账号审核"

try:
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass


@admin.register(User)
class UserAdmin(DjangoUserAdmin):
    change_form_template = "admin/users/user/change_form.html"
    change_list_template = "admin/users/user/change_list.html"
    list_display = (
        "identity_code",
        "username_display",
        "tel",
        "email_display",
        "created_at",
        "approval_status",
        "go_to_review",
    )
    list_display_links = ("identity_code",)
    list_filter = ()
    search_fields = ("identity_code", "username", "tel")
    ordering = ("-created_at",)
    list_per_page = 20

    fieldsets = (
        ("教师账号信息", {"fields": ("identity_code", "username", "email", "password_mask", "tel", "avatar")}),
        ("审核状态", {"fields": ("review_action_panel", "is_active")}),
        ("系统信息", {"fields": ("last_login_display",), "classes": ("collapse",)}),
    )
    readonly_fields = (
        "last_login",
        "last_login_display",
        "review_action_panel",
        "password_mask",
    )
    field_label_map = {
        "username": "用户名",
        "email": "邮箱",
        "password": "密码",
        "identity_code": "身份编号",
        "tel": "手机号",
        "avatar": "头像",
        "is_approved": "审核通过",
        "is_active": "是否启用",
        "last_login_display": "最后登录时间",
    }

    def get_queryset(self, request: HttpRequest):
        # Only show teacher accounts in the review table.
        return super().get_queryset(request).filter(role="teacher")

    def has_add_permission(self, request: HttpRequest):
        return False

    def has_delete_permission(self, request: HttpRequest, obj=None):
        return False

    def changelist_view(self, request: HttpRequest, extra_context=None):
        extra_context = extra_context or {}
        extra_context["title"] = "教师账号审核"
        return super().changelist_view(request, extra_context=extra_context)

    def get_form(self, request: HttpRequest, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        for field_name, label in self.field_label_map.items():
            if field_name in form.base_fields:
                form.base_fields[field_name].label = label
        # Remove default Django help texts in review page.
        for field_name in ("username", "email", "is_active"):
            if field_name in form.base_fields:
                form.base_fields[field_name].help_text = ""
        return form

    @admin.display(description="密码")
    def password_mask(self, obj: User):
        change_url = reverse("admin:users_user_change", args=[obj.pk])
        reset_url = change_url.replace("change/", "password/")
        return format_html(
            '<span class="password-mask">********</span>'
            '<a class="password-reset-link" href="{}">重置密码</a>',
            reset_url,
        )

    @admin.display(description="用户名")
    def username_display(self, obj: User):
        return obj.username

    @admin.display(description="邮箱")
    def email_display(self, obj: User):
        return obj.email or "-"

    @admin.display(description="最后登录时间")
    def last_login_display(self, obj: User):
        return obj.last_login or "-"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/approve/",
                self.admin_site.admin_view(self.approve_teacher),
                name="users_user_approve",
            ),
            path(
                "<path:object_id>/reject/",
                self.admin_site.admin_view(self.reject_teacher),
                name="users_user_reject",
            ),
        ]
        return custom_urls + urls

    def _redirect_after_review(self, request: HttpRequest, object_id: str):
        next_url = request.GET.get("next")
        if next_url:
            return HttpResponseRedirect(next_url)
        change_url = reverse("admin:users_user_change", args=[object_id])
        return HttpResponseRedirect(change_url)

    def approve_teacher(self, request: HttpRequest, object_id: str):
        user = get_object_or_404(User, pk=object_id, role="teacher")
        user.is_approved = True
        user.save(update_fields=["is_approved", "updated_at"])
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "approved": True, "message": "已通过该教师账号"})
        messages.success(request, f"已同意教师账号：{user.identity_code}")
        return self._redirect_after_review(request, object_id)

    def reject_teacher(self, request: HttpRequest, object_id: str):
        user = get_object_or_404(User, pk=object_id, role="teacher")
        user.is_approved = False
        user.save(update_fields=["is_approved", "updated_at"])
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "approved": False, "message": "已拒绝该教师账号"})
        messages.warning(request, f"已拒绝教师账号：{user.identity_code}")
        return self._redirect_after_review(request, object_id)

    @admin.display(description="前往审核")
    def go_to_review(self, obj: User):
        review_url = reverse("admin:users_user_change", args=[obj.pk])
        return format_html('<a href="{}">前往审核</a>', review_url)

    @admin.display(description="审核操作")
    def review_action_panel(self, obj: User):
        approve_url = reverse("admin:users_user_approve", args=[obj.pk]) + f"?next={reverse('admin:users_user_change', args=[obj.pk])}"
        reject_url = reverse("admin:users_user_reject", args=[obj.pk]) + f"?next={reverse('admin:users_user_change', args=[obj.pk])}"
        return format_html(
            '<a class="review-action-link" data-action="approve" href="{}" style="margin-right:10px;">通过</a>'
            '<a class="review-action-link" data-action="reject" href="{}">拒绝</a>'
            '<span id="review-action-feedback" style="margin-left:10px;color:#666;"></span>',
            approve_url,
            reject_url,
        )

    @admin.display(description="审核状态")
    def approval_status(self, obj: User):
        return "√" if obj.is_approved else "×"


class CatalogImportMixin:
    @staticmethod
    def _norm_header(text: str) -> str:
        return (text or "").strip().lower().replace(" ", "")

    @staticmethod
    def _pick_value(row_data, keys):
        for key in keys:
            value = row_data.get(CatalogImportMixin._norm_header(key))
            if value is not None and str(value).strip():
                return str(value).strip()
        return ""

    def _render_popup(self, request: HttpRequest, title: str, lines: list[str], level: str = "error"):
        return self.changelist_view(
            request,
            extra_context={
                "import_popup": {
                    "title": title,
                    "lines": lines,
                    "level": level,
                }
            },
        )

    def _load_rows(self, file):
        if not file:
            return None, "请先选择上传文件"
        if not file.name.lower().endswith(".xlsx"):
            return None, "仅支持 .xlsx 文件"
        try:
            workbook = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
        except Exception:
            return None, "模板解析失败，请确认文件内容正确"

        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if len(all_rows) < 2:
            return None, "模板中没有可导入的数据"

        headers = [self._norm_header(str(item) if item is not None else "") for item in all_rows[0]]
        row_items = []
        for idx, row in enumerate(all_rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_items.append((idx, dict(zip(headers, row))))
        if not row_items:
            return None, "模板中没有可导入的数据"
        return row_items, ""


@admin.register(MajorCatalog)
class MajorCatalogAdmin(CatalogImportMixin, admin.ModelAdmin):
    change_list_template = "admin/users/majorcatalog/change_list.html"
    list_display = ("name", "code")
    search_fields = ("name", "code")
    ordering = ("code",)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("upload/", self.admin_site.admin_view(self.upload_major_template), name="users_majorcatalog_upload"),
            path(
                "download-template/",
                self.admin_site.admin_view(self.download_major_template),
                name="users_majorcatalog_download_template",
            ),
        ]
        return custom_urls + urls

    def _get_template_path(self):
        return Path(settings.BASE_DIR) / "templates" / "data_template" / "major_upload_template.xlsx"

    def download_major_template(self, request: HttpRequest):
        template_path = self._get_template_path()
        if not template_path.exists():
            self.message_user(request, "专业模板不存在，请检查 data_template 目录", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:users_majorcatalog_changelist"))
        response = FileResponse(
            open(template_path, "rb"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="major_upload_template.xlsx"'
        return response

    def upload_major_template(self, request: HttpRequest):
        row_items, err = self._load_rows(request.FILES.get("file"))
        if err:
            return self._render_popup(request, "专业导入校验失败", [err], level="error")

        errors = []
        seen_codes = set()
        seen_names = set()
        normalized_rows = []
        for row_no, row_data in row_items:
            name = self._pick_value(row_data, ["专业", "专业名称", "major", "name"])
            code = self._pick_value(row_data, ["专业码", "专业代码", "majorcode", "major_code", "code"])
            row_errs = []
            if not name:
                row_errs.append("专业必填")
            if not code:
                row_errs.append("专业码必填")
            if code and code in seen_codes:
                row_errs.append(f"专业码重复({code})")
            if name and name in seen_names:
                row_errs.append(f"专业名称重复({name})")
            seen_codes.add(code)
            seen_names.add(name)
            if row_errs:
                errors.append(f"第{row_no}行：{'；'.join(row_errs)}")
            else:
                normalized_rows.append({"name": name, "code": code})

        if errors:
            summary = f"校验失败，共 {len(errors)} 条问题，未执行导入"
            return self._render_popup(request, "专业导入校验失败", [summary] + errors[:50], level="error")

        created_count = 0
        updated_count = 0
        for item in normalized_rows:
            _, created = MajorCatalog.objects.update_or_create(code=item["code"], defaults={"name": item["name"]})
            created_count += 1 if created else 0
            updated_count += 0 if created else 1
        return self._render_popup(
            request,
            "专业导入完成",
            [f"新增 {created_count} 条，更新 {updated_count} 条"],
            level="success",
        )


@admin.register(ClassCatalog)
class ClassCatalogAdmin(CatalogImportMixin, admin.ModelAdmin):
    change_list_template = "admin/users/classcatalog/change_list.html"
    list_display = ("name", "code", "major")
    search_fields = ("name", "code")
    list_select_related = ("major",)
    ordering = ("code",)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("upload/", self.admin_site.admin_view(self.upload_class_template), name="users_classcatalog_upload"),
            path(
                "download-template/",
                self.admin_site.admin_view(self.download_class_template),
                name="users_classcatalog_download_template",
            ),
        ]
        return custom_urls + urls

    def _get_template_path(self):
        return Path(settings.BASE_DIR) / "templates" / "data_template" / "class_upload_template.xlsx"

    def download_class_template(self, request: HttpRequest):
        template_path = self._get_template_path()
        if not template_path.exists():
            self.message_user(request, "班级模板不存在，请检查 data_template 目录", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:users_classcatalog_changelist"))
        response = FileResponse(
            open(template_path, "rb"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="class_upload_template.xlsx"'
        return response

    def upload_class_template(self, request: HttpRequest):
        row_items, err = self._load_rows(request.FILES.get("file"))
        if err:
            return self._render_popup(request, "班级导入校验失败", [err], level="error")

        major_by_name = {m.name: m for m in MajorCatalog.objects.all()}
        major_by_code = {m.code: m for m in MajorCatalog.objects.all()}
        errors = []
        seen_codes = set()
        normalized_rows = []
        for row_no, row_data in row_items:
            name = self._pick_value(row_data, ["班级", "班级名称", "classname", "class_name", "name"])
            code = self._pick_value(row_data, ["班级码", "班级代码", "classcode", "class_code", "code"])
            major_name = self._pick_value(row_data, ["专业", "专业名称"])
            major_code = self._pick_value(row_data, ["专业码", "专业代码", "majorcode", "major_code"])

            row_errs = []
            if not name:
                row_errs.append("班级名称必填")
            if not code:
                row_errs.append("班级码必填")
            if code and code in seen_codes:
                row_errs.append(f"班级码重复({code})")
            seen_codes.add(code)

            major = None
            if major_code:
                major = major_by_code.get(major_code)
                if major is None:
                    row_errs.append(f"专业码不存在({major_code})")
            if major is None and major_name:
                major = major_by_name.get(major_name)
                if major is None:
                    row_errs.append(f"专业名称不存在({major_name})")
            if row_errs:
                errors.append(f"第{row_no}行：{'；'.join(row_errs)}")
            else:
                normalized_rows.append({"name": name, "code": code, "major": major})

        if errors:
            summary = f"校验失败，共 {len(errors)} 条问题，未执行导入"
            return self._render_popup(request, "班级导入校验失败", [summary] + errors[:50], level="error")

        created_count = 0
        updated_count = 0
        for item in normalized_rows:
            _, created = ClassCatalog.objects.update_or_create(
                code=item["code"],
                defaults={"name": item["name"], "major": item["major"]},
            )
            created_count += 1 if created else 0
            updated_count += 0 if created else 1
        return self._render_popup(
            request,
            "班级导入完成",
            [f"新增 {created_count} 条，更新 {updated_count} 条"],
            level="success",
        )


@admin.register(MajorClassManagement)
class MajorClassManagementAdmin(admin.ModelAdmin):
    change_list_template = "admin/users/majorclassmanagement/change_list.html"

    def get_model_perms(self, request):
        # 仅用于自定义页面，不在左侧菜单单独显示
        return {}

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return MajorCatalog.objects.none()

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["title"] = "专业班级管理"
        extra_context["major_count"] = MajorCatalog.objects.count()
        extra_context["class_count"] = ClassCatalog.objects.count()
        extra_context["major_url"] = reverse("admin:users_majorcatalog_changelist")
        extra_context["class_url"] = reverse("admin:users_classcatalog_changelist")
        return super().changelist_view(request, extra_context=extra_context)


def custom_get_app_list(self, request, app_label=None):
    app_dict = self._build_app_dict(request, app_label)
    app_list = []

    users_app = app_dict.pop("users", None)
    if users_app:
        user_models = []
        catalog_models = []
        for model in users_app.get("models", []):
            object_name = model.get("object_name")
            if object_name in {"MajorCatalog", "ClassCatalog"}:
                renamed = dict(model)
                if object_name == "MajorCatalog":
                    renamed["name"] = "专业管理"
                if object_name == "ClassCatalog":
                    renamed["name"] = "班级管理"
                catalog_models.append(renamed)
            elif object_name != "MajorClassManagement":
                user_models.append(model)

        if user_models:
            users_entry = dict(users_app)
            users_entry["name"] = "Users"
            users_entry["models"] = sorted(user_models, key=lambda x: x["name"])
            app_list.append(users_entry)

        if catalog_models:
            catalog_entry = {
                "name": "专业班级管理",
                "app_label": "catalog_manage",
                "app_url": reverse("admin:users_majorcatalog_changelist"),
                "has_module_perms": True,
                "models": sorted(catalog_models, key=lambda x: x["name"]),
            }
            app_list.append(catalog_entry)

    for app in app_dict.values():
        app["models"].sort(key=lambda x: x["name"])
        app_list.append(app)

    app_list.sort(key=lambda x: x["name"].lower())
    return app_list


admin.site.get_app_list = MethodType(custom_get_app_list, admin.site)
