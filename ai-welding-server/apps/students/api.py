import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Optional
from uuid import uuid4

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import transaction
from django.http import FileResponse, JsonResponse
from ninja import File, Router, Schema, UploadedFile
from openpyxl import Workbook, load_workbook

from apps.users.models import ClassCatalog, MajorCatalog, Student
from core.auth import JWTAuth

router = Router(tags=["students"])
User = get_user_model()
logger = logging.getLogger(__name__)


class StudentOut(Schema):
    id: str
    identity_code: str
    username: str
    major: str
    major_code: str
    class_code: str
    class_name: str
    created_at: str


class StudentImportValidateOut(Schema):
    import_id: str
    total_rows: int
    valid_count: int
    invalid_count: int
    valid_rows: list[dict[str, Any]]
    invalid_rows: list[dict[str, Any]]
    can_import: bool
    message: str


class StudentImportCommitIn(Schema):
    import_id: str
    ignore_invalid: bool = False


class StudentImportCommitOut(Schema):
    message: str
    total_rows: int
    imported_count: int
    skipped_existing_count: int
    invalid_count: int


def _normalize_header(text: str) -> str:
    return (text or "").strip().lower().replace(" ", "")


def _pick_value(row: dict, keys: list[str]) -> str:
    for key in keys:
        if key in row and row[key] is not None:
            return str(row[key]).strip()
    return ""


def _build_unique_username(base_username: str, identity_code: str) -> str:
    candidate = (base_username or "").strip() or identity_code
    if not User.objects.filter(username=candidate).exists():
        return candidate

    fallback = f"{candidate}_{identity_code}"
    if len(fallback) > 150:
        fallback = fallback[:150]
    if not User.objects.filter(username=fallback).exists():
        return fallback

    return identity_code


def _parse_student_rows(file: UploadedFile):
    if not file.name.lower().endswith(".xlsx"):
        return None, "仅支持 .xlsx 格式"
    try:
        workbook = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
    except Exception:
        return None, "Excel 文件无法解析，请检查格式"

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return None, "Excel 至少需要 1 行数据"

    header_raw = rows[0]
    headers = [_normalize_header(str(item) if item is not None else "") for item in header_raw]
    row_dicts = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)]
    if not row_dicts:
        return None, "Excel 无有效数据行"
    return row_dicts, None


def _validate_student_rows(row_dicts: list[dict[str, Any]]):
    major_map = {item.name: item.code for item in MajorCatalog.objects.all()}
    class_map = {item.code: item for item in ClassCatalog.objects.select_related("major")}
    seen_identity_codes: set[str] = set()
    valid_rows: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, Any]] = []

    for index, row in enumerate(row_dicts, start=2):
        identity_code = _pick_value(row, ["学号", "student_id", "identity_code", "identitycode"])
        username = _pick_value(row, ["姓名", "name", "username"]) or identity_code
        major_name = _pick_value(row, ["专业", "专业名称", "major"])
        class_code = _pick_value(row, ["班级", "班级码", "班级code", "班级代码", "classcode", "class_code"])
        errors: list[str] = []

        if not identity_code:
            errors.append("学号必填")
        if not major_name:
            errors.append("专业必填")
        if identity_code:
            if identity_code in seen_identity_codes:
                errors.append(f"学号重复（{identity_code}）")
            seen_identity_codes.add(identity_code)

            existing_user = User.objects.filter(identity_code=identity_code).only("id", "role").first()
            if existing_user and existing_user.role != "student":
                errors.append(f"学号“{identity_code}”已被非学生账号占用")

        major_code = ""
        if major_name:
            major_code = major_map.get(major_name, "")
            if not major_code:
                errors.append(f"专业“{major_name}”未配置专业字典")

        class_name = ""
        if class_code:
            class_item = class_map.get(class_code)
            if not class_item:
                errors.append(f"班级码“{class_code}”未配置班级字典")
            else:
                class_name = class_item.name
                if class_item.major and major_name and class_item.major.name != major_name:
                    errors.append(f"班级码“{class_code}”所属专业与导入专业“{major_name}”不一致")

        normalized = {
            "row_no": index,
            "identity_code": identity_code,
            "username": username,
            "major": major_name,
            "major_code": major_code,
            "class_code": class_code,
            "class_name": class_name,
        }
        if errors:
            invalid_rows.append({**normalized, "error_message": "；".join(errors)})
        else:
            valid_rows.append(normalized)

    return valid_rows, invalid_rows


@router.get("/template", auth=JWTAuth())
def download_student_template(request):
    logger.info("Download student template requested user_id=%s", getattr(request.auth, "id", None))
    template_path = Path(settings.BASE_DIR) / "templates" / "data_template" / "student_upload_template.xlsx"
    if not template_path.exists():
        logger.error("Student template missing path=%s", template_path)
        return JsonResponse({"message": "模板文件不存在，请联系管理员"}, status=400)

    response = FileResponse(
        open(template_path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_upload_template.xlsx"'
    return response


@router.post("/import/validate", auth=JWTAuth(), response={200: StudentImportValidateOut, 400: dict, 403: dict})
def validate_students_import(request, file: UploadedFile = File(...)):
    if request.auth.role != "teacher":
        logger.warning("Validate students import forbidden user_id=%s role=%s", getattr(request.auth, "id", None), request.auth.role)
        return 403, {"message": "仅教师可执行导入"}

    logger.info("Validate students import start user_id=%s filename=%s", getattr(request.auth, "id", None), file.name)
    row_dicts, error = _parse_student_rows(file)
    if error:
        logger.warning("Validate students import failed user_id=%s reason=%s", getattr(request.auth, "id", None), error)
        return 400, {"message": error}

    valid_rows, invalid_rows = _validate_student_rows(row_dicts)
    import_id = str(uuid4())
    cache.set(
        f"student_import:{import_id}",
        {"valid_rows": valid_rows, "invalid_rows": invalid_rows, "total_rows": len(row_dicts)},
        timeout=1800,
    )
    logger.info(
        "Validate students import done user_id=%s total=%s valid=%s invalid=%s import_id=%s",
        getattr(request.auth, "id", None),
        len(row_dicts),
        len(valid_rows),
        len(invalid_rows),
        import_id,
    )

    if len(valid_rows) == len(row_dicts):
        message = "数据校验通过，可以继续导入"
    elif len(valid_rows) == 0:
        message = "数据有误，请先修正"
    else:
        message = "存在部分异常数据，可忽视异常继续导入正确数据"

    return {
        "import_id": import_id,
        "total_rows": len(row_dicts),
        "valid_count": len(valid_rows),
        "invalid_count": len(invalid_rows),
        "valid_rows": valid_rows[:200],
        "invalid_rows": invalid_rows[:200],
        "can_import": len(valid_rows) > 0,
        "message": message,
    }


@router.post("/import/commit", auth=JWTAuth(), response={200: StudentImportCommitOut, 400: dict, 403: dict})
def commit_students_import(request, payload: StudentImportCommitIn):
    if request.auth.role != "teacher":
        logger.warning("Commit students import forbidden user_id=%s role=%s", getattr(request.auth, "id", None), request.auth.role)
        return 403, {"message": "仅教师可执行导入"}

    logger.info("Commit students import start user_id=%s import_id=%s ignore_invalid=%s", getattr(request.auth, "id", None), payload.import_id, payload.ignore_invalid)
    cached = cache.get(f"student_import:{payload.import_id}")
    if not cached:
        logger.warning("Commit students import failed user_id=%s import_id=%s reason=session_expired", getattr(request.auth, "id", None), payload.import_id)
        return 400, {"message": "导入会话已过期，请重新上传文件"}

    valid_rows = cached.get("valid_rows", [])
    invalid_rows = cached.get("invalid_rows", [])
    total_rows = cached.get("total_rows", 0)
    if not valid_rows:
        logger.warning("Commit students import failed user_id=%s import_id=%s reason=no_valid_rows", getattr(request.auth, "id", None), payload.import_id)
        return 400, {"message": "数据有误，请先修正后再导入"}
    if invalid_rows and not payload.ignore_invalid:
        logger.warning("Commit students import blocked user_id=%s import_id=%s reason=invalid_rows_need_confirm", getattr(request.auth, "id", None), payload.import_id)
        return 400, {"message": "存在异常数据，请选择忽视异常后再继续导入"}

    imported_count = 0
    skipped_existing_count = 0

    with transaction.atomic():
        for row in valid_rows:
            identity_code = row["identity_code"]
            existing_user = User.objects.filter(identity_code=identity_code).only("id", "role").first()
            if existing_user:
                skipped_existing_count += 1
                continue

            user = User.objects.create_user(
                username=_build_unique_username(row["username"], identity_code),
                identity_code=identity_code,
                password=identity_code,
                role="student",
                is_approved=True,
            )
            Student.objects.create(
                user=user,
                major=row["major"],
                major_code=row["major_code"],
                class_name=row["class_name"],
                class_code=row["class_code"],
            )
            imported_count += 1

    cache.delete(f"student_import:{payload.import_id}")
    logger.info(
        "Commit students import done user_id=%s import_id=%s imported=%s skipped_existing=%s invalid=%s",
        getattr(request.auth, "id", None),
        payload.import_id,
        imported_count,
        skipped_existing_count,
        len(invalid_rows),
    )
    return {
        "message": "导入完成",
        "total_rows": total_rows,
        "imported_count": imported_count,
        "skipped_existing_count": skipped_existing_count,
        "invalid_count": len(invalid_rows),
    }


@router.get("/import/{import_id}/invalid-template", auth=JWTAuth())
def download_invalid_template(request, import_id: str):
    if request.auth.role != "teacher":
        logger.warning("Download invalid template forbidden user_id=%s role=%s", getattr(request.auth, "id", None), request.auth.role)
        return JsonResponse({"message": "仅教师可下载"}, status=403)

    logger.info("Download invalid template requested user_id=%s import_id=%s", getattr(request.auth, "id", None), import_id)
    cached = cache.get(f"student_import:{import_id}")
    if not cached:
        logger.warning("Download invalid template failed import_id=%s reason=session_expired", import_id)
        return JsonResponse({"message": "导入会话已过期，请重新上传文件"}, status=400)

    invalid_rows = cached.get("invalid_rows", [])
    if not invalid_rows:
        logger.warning("Download invalid template failed import_id=%s reason=no_invalid_rows", import_id)
        return JsonResponse({"message": "当前没有异常数据"}, status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "invalid_students"
    ws.append(["学号", "姓名", "专业", "班级", "异常信息"])
    for row in invalid_rows:
        ws.append(
            [
                row.get("identity_code", ""),
                row.get("username", ""),
                row.get("major", ""),
                row.get("class_code", ""),
                row.get("error_message", ""),
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = FileResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_invalid_rows.xlsx"'
    return response


@router.get("", auth=JWTAuth(), response={200: list[StudentOut], 403: dict})
def list_students(
    request,
    identity_code: Optional[str] = None,
    major_code: Optional[str] = None,
    class_code: Optional[str] = None,
    class_name: Optional[str] = None,
    name: Optional[str] = None,
):
    if request.auth.role != "teacher":
        logger.warning("List students forbidden user_id=%s role=%s", getattr(request.auth, "id", None), request.auth.role)
        return 403, {"message": "仅教师可查看学生信息"}

    logger.info(
        "List students requested user_id=%s filters(identity_code=%s, major_code=%s, class_code=%s, class_name=%s, name=%s)",
        getattr(request.auth, "id", None),
        identity_code,
        major_code,
        class_code,
        class_name,
        name,
    )
    queryset = Student.objects.select_related("user").filter(user__role="student")
    if identity_code:
        queryset = queryset.filter(user__identity_code__icontains=identity_code.strip())
    if major_code:
        queryset = queryset.filter(major_code__icontains=major_code.strip())
    if class_code:
        queryset = queryset.filter(class_code__icontains=class_code.strip())
    if class_name:
        queryset = queryset.filter(class_name__icontains=class_name.strip())
    if name:
        queryset = queryset.filter(user__username__icontains=name.strip())

    return [
        {
            "id": str(student.id),
            "identity_code": student.user.identity_code,
            "username": student.user.username,
            "major": student.major,
            "major_code": student.major_code,
            "class_code": student.class_code,
            "class_name": student.class_name,
            "created_at": student.user.created_at.isoformat(),
        }
        for student in queryset
    ]
