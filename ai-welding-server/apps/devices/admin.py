from io import BytesIO
from datetime import datetime
from django.contrib import admin, messages
from django.http import FileResponse, HttpRequest, HttpResponseRedirect
from django.urls import path, reverse
from openpyxl import Workbook, load_workbook

from .models import Device


@admin.register(Device)
class DeviceAdmin(admin.ModelAdmin):
    change_list_template = "admin/devices/device/change_list.html"
    list_display = ("device_code", "status", "classroom", "start_time", "end_time", "updated_at")
    search_fields = ("device_code", "classroom")
    list_filter = ("status", "classroom")
    ordering = ("device_code",)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("upload/", self.admin_site.admin_view(self.upload_template), name="devices_device_upload"),
            path(
                "download-template/",
                self.admin_site.admin_view(self.download_template),
                name="devices_device_download_template",
            ),
        ]
        return custom_urls + urls

    @staticmethod
    def _norm_header(text: str) -> str:
        return (text or "").strip().lower().replace(" ", "")

    @staticmethod
    def _pick_value(row_data: dict, keys: list[str]) -> str:
        for key in keys:
            value = row_data.get(DeviceAdmin._norm_header(key))
            if value is not None and str(value).strip():
                return str(value).strip()
        return ""

    @staticmethod
    def _to_status(value: str) -> str:
        mapping = {
            "使用中": Device.STATUS_IN_USE,
            "空闲": Device.STATUS_IDLE,
            "维护中": Device.STATUS_MAINTAINING,
            "in_use": Device.STATUS_IN_USE,
            "idle": Device.STATUS_IDLE,
            "maintaining": Device.STATUS_MAINTAINING,
        }
        return mapping.get((value or "").strip(), "")

    @staticmethod
    def _to_datetime(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def download_template(self, request: HttpRequest):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "devices"
        sheet.append(["设备编号", "状态", "教室", "开始时间", "结束时间"])
        sheet.append(["EQ-001", "使用中", "A室-01", "2026-04-28 08:00:00", "2026-04-28 18:00:00"])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = FileResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="device_upload_template.xlsx"'
        return response

    def upload_template(self, request: HttpRequest):
        file = request.FILES.get("file")
        if not file:
            self.message_user(request, "请先选择上传文件", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:devices_device_changelist"))
        if not file.name.lower().endswith(".xlsx"):
            self.message_user(request, "仅支持 .xlsx 文件", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:devices_device_changelist"))

        try:
            workbook = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
        except Exception:
            self.message_user(request, "模板解析失败，请确认内容正确", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:devices_device_changelist"))

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            self.message_user(request, "模板中没有可导入数据", level=messages.ERROR)
            return HttpResponseRedirect(reverse("admin:devices_device_changelist"))

        headers = [self._norm_header(str(item) if item is not None else "") for item in rows[0]]
        created_count = 0
        updated_count = 0
        errors: list[str] = []

        for idx, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = dict(zip(headers, row))
            code = self._pick_value(row_data, ["设备编号", "device_code", "code"])
            status_text = self._pick_value(row_data, ["状态", "status"])
            classroom = self._pick_value(row_data, ["教室", "classroom"])
            start_time_raw = row_data.get(self._norm_header("开始时间")) or row_data.get(self._norm_header("start_time"))
            end_time_raw = row_data.get(self._norm_header("结束时间")) or row_data.get(self._norm_header("end_time"))
            status = self._to_status(status_text)
            start_time = self._to_datetime(start_time_raw)
            end_time = self._to_datetime(end_time_raw)

            row_errors = []
            if not code:
                row_errors.append("设备编号必填")
            if not status:
                row_errors.append("状态必须是：使用中/空闲/维护中")
            if not classroom:
                row_errors.append("教室必填")
            if start_time_raw and start_time is None:
                row_errors.append("开始时间格式错误，支持 YYYY-MM-DD HH:MM[:SS]")
            if end_time_raw and end_time is None:
                row_errors.append("结束时间格式错误，支持 YYYY-MM-DD HH:MM[:SS]")
            if start_time and end_time and start_time > end_time:
                row_errors.append("开始时间不能晚于结束时间")
            if row_errors:
                errors.append(f"第{idx}行：{'；'.join(row_errors)}")
                continue

            _, created = Device.objects.update_or_create(
                device_code=code,
                defaults={"status": status, "classroom": classroom, "start_time": start_time, "end_time": end_time},
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        if errors:
            for msg in errors[:20]:
                self.message_user(request, msg, level=messages.ERROR)
            self.message_user(request, f"导入结束：新增 {created_count}，更新 {updated_count}，错误 {len(errors)}", level=messages.WARNING)
        else:
            self.message_user(request, f"导入完成：新增 {created_count}，更新 {updated_count}", level=messages.SUCCESS)
        return HttpResponseRedirect(reverse("admin:devices_device_changelist"))
